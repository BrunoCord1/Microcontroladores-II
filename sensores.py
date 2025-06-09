import serial
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import subprocess
import matplotlib.pyplot as plt

# === CONFIGURACIÓN DEL PUERTO SERIAL ===
puerto_serial = serial.Serial("COM3", 9600, timeout=1)  # Ajusta el COM según tu PC

# === ARCHIVO EXCEL ===
nombre_archivo = "sensores.xlsx"

# Crear si no existe
if not os.path.exists(nombre_archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mediciones"
    encabezados = ["Fecha/Hora", "Humedad Suelo", "Temp LM35 (°C)", "Distancia (cm)", "Temp DHT22 (°C)", "Humedad DHT22 (%)"]
    ws.append(encabezados)
    wb.save(nombre_archivo)

# Cargar archivo existente
wb = load_workbook(nombre_archivo)
ws = wb.active

# === VENTANA PRINCIPAL ===
ventana = tk.Tk()
ventana.title("Monitor de Sensores")
ventana.geometry("500x400")
ventana.configure(bg="#f0f0f0")

style = ttk.Style()
style.configure("TLabel", font=("Segoe UI", 10))
style.configure("TButton", font=("Segoe UI", 10, "bold"))

# Título
titulo = ttk.Label(ventana, text="Monitor en Tiempo Real", font=("Segoe UI", 14, "bold"))
titulo.grid(row=0, column=0, columnspan=2, pady=10)

# Variables para datos
valores = {
    "Humedad del suelo": tk.StringVar(),
    "Temp LM35 (°C)": tk.StringVar(),
    "Distancia (cm)": tk.StringVar(),
    "Temp Aire DHT22 (°C)": tk.StringVar(),
    "Humedad Aire DHT22 (%)": tk.StringVar()
}

fila = 1
for nombre, var in valores.items():
    ttk.Label(ventana, text=nombre + ":").grid(row=fila, column=0, sticky="e", padx=10, pady=5)
    ttk.Label(ventana, textvariable=var, foreground="blue").grid(row=fila, column=1, sticky="w")
    fila += 1

# === FUNCIONES ===

def actualizar_datos():
    if puerto_serial.in_waiting:
        try:
            linea = puerto_serial.readline().decode("utf-8").strip()
            partes = linea.split(",")
            if len(partes) == 5:
                valores["Humedad del suelo"].set(partes[0])
                valores["Temp LM35 (°C)"].set(partes[1])
                valores["Distancia (cm)"].set(partes[2])
                valores["Temp Aire DHT22 (°C)"].set(partes[3])
                valores["Humedad Aire DHT22 (%)"].set(partes[4])

                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([
                    timestamp,
                    int(partes[0]),
                    float(partes[1]),
                    float(partes[2]),
                    float(partes[3]),
                    float(partes[4])
                ])
                wb.save(nombre_archivo)
        except Exception as e:
            print("Error:", e)

    ventana.after(1000, actualizar_datos)

def abrir_excel():
    try:
        subprocess.run(["start", nombre_archivo], shell=True)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")

def mostrar_graficas():
    try:
        wb_local = load_workbook(nombre_archivo)
        ws_local = wb_local.active

        fechas = []
        humedad_suelo = []
        temp_lm35 = []
        distancia = []
        temp_dht22 = []
        humedad_dht22 = []

        for row in ws_local.iter_rows(min_row=2, values_only=True):
            fechas.append(row[0])
            humedad_suelo.append(row[1])
            temp_lm35.append(row[2])
            distancia.append(row[3])
            temp_dht22.append(row[4])
            humedad_dht22.append(row[5])

        fig, axs = plt.subplots(3, 2, figsize=(12, 8))
        fig.suptitle("Gráficas de Sensores", fontsize=16)

        axs[0, 0].plot(fechas, humedad_suelo, color='green')
        axs[0, 0].set_title("Humedad del Suelo")
        axs[0, 0].tick_params(labelrotation=45)

        axs[0, 1].plot(fechas, temp_lm35, color='red')
        axs[0, 1].set_title("Temp LM35 (°C)")
        axs[0, 1].tick_params(labelrotation=45)

        axs[1, 0].plot(fechas, distancia, color='blue')
        axs[1, 0].set_title("Distancia (cm)")
        axs[1, 0].tick_params(labelrotation=45)

        axs[1, 1].plot(fechas, temp_dht22, color='orange')
        axs[1, 1].set_title("Temp DHT22 (°C)")
        axs[1, 1].tick_params(labelrotation=45)

        axs[2, 0].plot(fechas, humedad_dht22, color='purple')
        axs[2, 0].set_title("Humedad DHT22 (%)")
        axs[2, 0].tick_params(labelrotation=45)

        fig.delaxes(axs[2, 1])

        plt.tight_layout(rect=[0, 0, 1, 0.95])
        plt.show()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron generar las gráficas:\n{e}")

# === BOTONES ===
ttk.Button(ventana, text="Abrir Excel", command=abrir_excel).grid(row=fila, column=0, pady=20)
ttk.Button(ventana, text="Ver Gráficas", command=mostrar_graficas).grid(row=fila, column=1, pady=20)
fila += 1
ttk.Button(ventana, text="Salir", command=ventana.destroy).grid(row=fila, column=0, columnspan=2, pady=5)

# === INICIAR LECTURA Y VENTANA ===
actualizar_datos()
ventana.mainloop()
